#!/usr/bin/env python3
# coding=utf-8

import openpyxl
# 创建一个工作簿
# wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单
# wb.create_sheet('test_case')
# 保存为一个xlsx格式的文件
# wb.save('D:\cases.xlsx')
# 读取excel中的数据
# 第一步：打开工作簿
wb = openpyxl.load_workbook('D:\1621577983.csv')
# 第二步：选取表单
sh = wb['test_case']
# 第三步：读取数据
# 参数 row:行  column：列
ce = sh.cell(row=1, column=1)   # 读取第一行，第一列的数据
print(ce.value)
# 按行读取数据 list(sh.rows)
print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据
for cases in list(sh.rows)[1:]:
    case_id = cases[0].value
    case_excepted = cases[1].value
    case_data = cases[2].value
    print(case_excepted, case_data)
# 关闭工作薄
wb.close()