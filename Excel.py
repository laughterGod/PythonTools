# coding=utf-8
import openpyxl

wb = openpyxl.load_workbook('D:\\pythonFiles\\a.xlsx')
sheet = wb.active

with open('example_1.txt', 'w') as f1:
    for i in range(1, sheet.max_row + 1):
        f1.write(str(sheet.cell(row=i, column=1).value))
f1.close()

with open('example_2.txt', 'w') as f2:
    for i in range(1, sheet.max_row + 1):
        f2.write(str(sheet.cell(row=i, column=2).value))
f2.close()

with open('example_3.txt', 'w') as f3:
    for i in range(1, sheet.max_row + 1):
        f3.write(str(sheet.cell(row=i, column=3).value))
f3.close()